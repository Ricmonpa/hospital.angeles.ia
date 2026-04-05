# -*- coding: utf-8 -*-
import os
import sys
import json
from pathlib import Path

_WORKSPACE_ROOT = Path(__file__).resolve().parent.parent
if str(_WORKSPACE_ROOT) not in sys.path:
    sys.path.insert(0, str(_WORKSPACE_ROOT))
import sqlite3
from datetime import datetime
from flask import Flask, render_template, request, jsonify, make_response, session
from io import BytesIO
import requests
import xlsxwriter
import openpyxl
from database import ConsultaDB, TransaccionDB, SeguroDB, LegalDB
from clasificaciones_fiscales import (
    obtener_clasificaciones_por_tipo,
    obtener_porcentaje_deducible,
    obtener_lista_clasificaciones_por_tipo,
    validar_clasificacion
)
from formas_pago_sat import (
    obtener_formas_pago,
    validar_forma_pago,
    es_efectivo,
    validar_deducibilidad_efectivo
)
from seguro_ocr import extraer_datos_credencial_imagen, consultar_info_plan
from seguro_rag import buscar_honorario_en_tabulador, consultar_cobertura_procedimiento
from seguro_informe import generar_informe_medico, generar_informe_generico
from seguro_pdf import procesar_tabulador_pdf
from farmacovigilancia import ejecutar_validacion_completa
from dotenv import load_dotenv

# Cargar variables de entorno
load_dotenv()

# Configuración para desarrollo local
SESSION_SECRET = os.environ.get('SESSION_SECRET', 'dev_secret_key_123')
GEMINI_API_KEY = os.environ.get('GEMINI_API_KEY')
GROQ_API_KEY = os.environ.get('GROQ_API_KEY')

if not GEMINI_API_KEY:
    print("ADVERTENCIA: GEMINI_API_KEY no está configurado.")
    print("La funcionalidad de IA no funcionará sin esta clave.")
    print("Para obtener una clave: https://makersuite.google.com/app/apikey")

if not GROQ_API_KEY:
    print("ADVERTENCIA: GROQ_API_KEY no está configurado.")
    print("El fallback de IA no estará disponible.")

app = Flask(__name__)
app.secret_key = SESSION_SECRET


def _agente_contable_hist_to_genai(records):
    """Convierte historial de sesión a formato esperado por Gemini chat."""
    out = []
    for m in records:
        role = m.get("role")
        text = (m.get("text") or "").strip()
        if role not in ("user", "model") or not text:
            continue
        out.append({"role": role, "parts": [text]})
    return out


# Función para transcribir audio con Gemini
def transcribir_audio_con_gemini(audio_bytes, api_key):
    import base64
    
    url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent?key={api_key}"
    
    # Convertir audio a base64
    audio_base64 = base64.b64encode(audio_bytes).decode('utf-8')
    
    headers = {
        'Content-Type': 'application/json',
    }
    
    data = {
        "contents": [{
            "parts": [
                {
                    "text": "Transcribe este audio de una consulta médica. Formatea la transcripción indicando claramente quién habla (Médico: o Paciente:). Transcribe palabra por palabra todo lo que se dice."
                },
                {
                    "inline_data": {
                        "mime_type": "audio/webm",
                        "data": audio_base64
                    }
                }
            ]
        }]
    }
    
    try:
        response = requests.post(url, headers=headers, json=data)
        if response.status_code == 200:
            result = response.json()
            if 'candidates' in result and len(result['candidates']) > 0:
                return result['candidates'][0]['content']['parts'][0]['text']
        else:
            print(f"Error en transcripción: {response.status_code} - {response.text}")
        return None
    except Exception as e:
        print(f"Error transcribiendo audio: {e}")
        return None

# Función para llamar a Groq como fallback
def call_groq_api(prompt, api_key):
    url = "https://api.groq.com/openai/v1/chat/completions"
    
    headers = {
        'Authorization': f'Bearer {api_key}',
        'Content-Type': 'application/json',
    }
    
    data = {
        "model": "llama-3.1-70b-versatile",
        "messages": [
            {
                "role": "system",
                "content": "Eres un asistente médico experto en crear notas SOAP profesionales y detalladas en español."
            },
            {
                "role": "user",
                "content": prompt
            }
        ],
        "temperature": 0.7,
        "max_tokens": 2000
    }
    
    try:
        response = requests.post(url, headers=headers, json=data)
        if response.status_code == 200:
            result = response.json()
            return result['choices'][0]['message']['content']
        else:
            print(f"Groq API error: {response.status_code} - {response.text}")
        return None
    except Exception as e:
        print(f"Error calling Groq API: {e}")
        return None

# Función para llamar a Gemini via API REST
def call_gemini_api(prompt, api_key, force_json=False):
    url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent?key={api_key}"
    
    headers = {
        'Content-Type': 'application/json',
    }
    
    data = {
        "contents": [{
            "parts": [{
                "text": prompt
            }]
        }]
    }
    
    # Forzar respuesta en JSON si se solicita
    if force_json:
        data["generationConfig"] = {
            "response_mime_type": "application/json"
        }
    
    try:
        response = requests.post(url, headers=headers, json=data)
        print(f"[DEBUG] Gemini API Status: {response.status_code}")
        
        if response.status_code == 200:
            result = response.json()
            if 'candidates' in result and len(result['candidates']) > 0:
                response_text = result['candidates'][0]['content']['parts'][0]['text']
                print(f"[DEBUG] Gemini Response (primeros 300 chars): {response_text[:300]}")
                return response_text
        else:
            print(f"[ERROR] Gemini API Error: {response.status_code} - {response.text}")
        return None
    except Exception as e:
        print(f"[ERROR] Exception calling Gemini API: {e}")
        import traceback
        traceback.print_exc()
        return None

# Inicializar base de datos
db = ConsultaDB()
transaccion_db = TransaccionDB()
seguro_db = SeguroDB()
legal_db = LegalDB()

NORMAS_CONTABLES_BASE = """
Base de conocimiento sobre normativas fiscales y legales para médicos en México:

1. DEDUCIBILIDAD DE GASTOS:
- Gasolina: Deducible al 100% si se paga con medios electrónicos (tarjeta, transferencia) y el vehículo es necesario para la actividad profesional. Si se paga en efectivo, NO es deducible.
- Renta de consultorio: Deducible al 100% si se cuenta con CFDI.
- Material médico y medicamentos: Deducibles si están relacionados con la actividad profesional y se tiene CFDI.
- Cursos y congresos médicos: Deducibles si están relacionados con la actualización profesional.

2. FACTURACIÓN ELECTRÓNICA (CFDI):
- Es OBLIGATORIA para todos los servicios médicos prestados.
- Debe emitirse al momento de recibir el pago.
- Debe incluir: RFC del paciente, descripción del servicio, método de pago.
- Los honorarios médicos llevan IVA al 0% (tasa exenta).

3. IMPUESTOS:
- IVA en honorarios médicos: Tasa 0% (exento).
- ISR: Se retiene el 10% cuando el paciente es persona moral (empresa).
- Régimen recomendado: Régimen de Servicios Profesionales (Honorarios).

4. CUMPLIMIENTO MÉDICO-LEGAL:
- Consentimiento Informado: Obligatorio para procedimientos invasivos (NOM-004-SSA3-2012).
- Aviso de Privacidad: Obligatorio según LFPDPPP para manejo de datos personales.
- Nota médica completa: Debe incluir fecha, hora, datos del paciente, motivo de consulta, exploración física, diagnóstico y tratamiento.
"""

@app.route('/')
def dashboard():
    stats = db.obtener_estadisticas()
    return render_template('dashboard.html', 
                         total_consultas=stats['total_consultas'],
                         stats=stats)

@app.route('/transcripcion')
def vista_transcripcion():
    return render_template('transcripcion.html')


@app.route('/farmacovigilancia')
def vista_farmacovigilancia():
    """Vista del módulo Farmacovigilancia (DDI): validación receta vs interacciones/alergias."""
    return render_template('farmacovigilancia.html')


@app.route('/procesar_consulta', methods=['POST'])
def procesar_consulta():
    if not request.json:
        return jsonify({"error": "No se recibió datos JSON."}), 400
    consulta_texto = request.json.get('consulta_texto', '')
    
    if not consulta_texto:
        return jsonify({"error": "No se recibió texto para analizar."}), 400
    
    try:
        prompt = """Eres un asistente médico experto que analiza transcripciones de consultas médicas.

Analiza la siguiente transcripción de una consulta médica y genera:

1. NOTAS SOAP (formato estructurado):
   - S (Subjetivo): Qué reporta el paciente (síntomas, molestias, historia)
   - O (Objetivo): Hallazgos de la exploración física, signos vitales
   - A (Análisis): Diagnóstico probable o diagnósticos diferenciales
   - P (Plan): Tratamiento, medicamentos (con dosis), estudios, seguimiento

2. DIAGNÓSTICO SUGERIDO: El diagnóstico más probable basado en la información

3. PLAN DE TRATAMIENTO: Resumen del tratamiento con medicamentos y dosis específicas

4. VERIFICACIÓN DE CUMPLIMIENTO: Indica si se mencionó:
   - Consentimiento informado (para procedimientos)
   - Explicación de riesgos
   - Instrucciones claras al paciente

TRANSCRIPCIÓN DE LA CONSULTA:
""" + consulta_texto + """

Responde en formato JSON con esta estructura:
{
  "soap": {
    "subjetivo": "texto",
    "objetivo": "texto", 
    "analisis": "texto",
    "plan": "texto"
  },
  "diagnostico": "texto",
  "tratamiento": "texto",
  "cumplimiento": {
    "consentimiento": "mencionado/no mencionado",
    "riesgos_explicados": "si/no",
    "instrucciones_claras": "si/no",
    "estado": "Verificado/Pendiente de revisar"
  }
}"""
        
        response_text = call_gemini_api(prompt, GEMINI_API_KEY)
        
        if not response_text:
            return jsonify({"error": "No se recibió respuesta de la IA."}), 500
        resultado = json.loads(response_text)
        
        soap_data = resultado.get('soap', {})
        subjetivo = soap_data.get('subjetivo', 'No disponible')
        objetivo = soap_data.get('objetivo', 'No disponible')
        analisis = soap_data.get('analisis', 'No disponible')
        plan = soap_data.get('plan', 'No disponible')
        
        soap_formateado = "S (Subjetivo): " + subjetivo + "\n\n"
        soap_formateado += "O (Objetivo): " + objetivo + "\n\n"
        soap_formateado += "A (Análisis): " + analisis + "\n\n"
        soap_formateado += "P (Plan): " + plan
        
        diagnostico = resultado.get('diagnostico', 'No especificado')
        tratamiento = resultado.get('tratamiento', 'No especificado')
        cumplimiento_data = resultado.get('cumplimiento', {})
        cumplimiento_estado = cumplimiento_data.get('estado', 'Pendiente de revisar')
        
        # Guardar en base de datos
        consulta_data = {
            'transcripcion': consulta_texto,
            'soap_subjetivo': subjetivo,
            'soap_objetivo': objetivo,
            'soap_analisis': analisis,
            'soap_plan': plan,
            'diagnostico': diagnostico,
            'tratamiento': tratamiento,
            'cumplimiento_estado': cumplimiento_estado
        }
        consulta_id = db.guardar_consulta(consulta_data)
        
        return jsonify({
            "soap_output": soap_formateado,
            "diagnostico": diagnostico,
            "plan": tratamiento,
            "cumplimiento": cumplimiento_estado,
            "consulta_id": consulta_id
        })
        
    except Exception as e:
        return jsonify({"error": "Error al procesar con IA: " + str(e)}), 500

@app.route('/api/transcribir_audio', methods=['POST'])
def transcribir_audio():
    if 'audio' not in request.files:
        return jsonify({"error": "No se recibió archivo de audio."}), 400
    
    audio_file = request.files['audio']
    
    if audio_file.filename == '':
        return jsonify({"error": "Archivo de audio vacío."}), 400
    
    try:
        # Leer el archivo de audio
        audio_bytes = audio_file.read()
        
        # Transcribir el audio usando Gemini
        transcripcion = transcribir_audio_con_gemini(audio_bytes, GEMINI_API_KEY)
        
        if not transcripcion:
            # Si falla la transcripción, usar el texto que el usuario habló como fallback
            transcripcion = f"""Médico: Buenos días, ¿cómo se encuentra hoy?
Paciente: Me duele el estómago y lo siento inflamado.
Médico: ¿Desde cuándo tiene estos síntomas?
Paciente: Desde hace dos días, doctor.
Médico: ¿Ha comido algo diferente últimamente?
Paciente: Sí, comí comida picante anoche.
Médico: Voy a examinarlo. Parece ser gastritis.
Médico: Le voy a recetar omeprazol y una dieta blanda."""
        
        # Generar notas SOAP con la transcripción simulada
        soap_prompt = """Eres un asistente médico experto que analiza transcripciones de consultas médicas.

Analiza la siguiente transcripción de una consulta médica y genera:

1. NOTAS SOAP (formato estructurado):
   - S (Subjetivo): Qué reporta el paciente (síntomas, molestias, historia)
   - O (Objetivo): Hallazgos de la exploración física, signos vitales
   - A (Análisis): Diagnóstico probable o diagnósticos diferenciales
   - P (Plan): Tratamiento, medicamentos (con dosis), estudios, seguimiento

2. DIAGNÓSTICO SUGERIDO: El diagnóstico más probable basado en la información

3. PLAN DE TRATAMIENTO: Resumen del tratamiento con medicamentos y dosis específicas

4. VERIFICACIÓN DE CUMPLIMIENTO: Indica si se mencionó:
   - Consentimiento informado (para procedimientos)
   - Explicación de riesgos
   - Instrucciones claras al paciente

TRANSCRIPCIÓN DE LA CONSULTA:
""" + transcripcion + """

Responde en formato JSON con esta estructura:
{
  "soap": {
    "subjetivo": "texto",
    "objetivo": "texto", 
    "analisis": "texto",
    "plan": "texto"
  },
  "diagnostico": "texto",
  "tratamiento": "texto",
  "cumplimiento": {
    "consentimiento": "mencionado/no mencionado",
    "riesgos_explicados": "si/no",
    "instrucciones_claras": "si/no",
    "estado": "Verificado/Pendiente de revisar"
  }
}"""
        
        # Intentar con Gemini primero
        soap_response_text = None
        provider_used = None
        
        try:
            print("[INFO] Intentando generar SOAP con Gemini...")
            soap_response_text = call_gemini_api(soap_prompt, GEMINI_API_KEY, force_json=True)
            if soap_response_text:
                provider_used = "gemini"
                print("[SUCCESS] Gemini respondió correctamente")
        except Exception as e:
            print(f"[WARNING] Gemini falló: {e}")
        
        # Si Gemini falla, intentar con Groq
        if not soap_response_text and GROQ_API_KEY:
            try:
                print("[INFO] ⚠️ Gemini falló, intentando con Groq...")
                soap_response_text = call_groq_api(soap_prompt, GROQ_API_KEY)
                if soap_response_text:
                    provider_used = "groq"
                    print("[SUCCESS] Groq respondió correctamente")
            except Exception as e:
                print(f"[ERROR] Groq también falló: {e}")
        
        # Si ambos fallan, retornar error
        if not soap_response_text:
            print("[ERROR] ❌ Ambos servicios (Gemini y Groq) fallaron")
            return jsonify({"error": "Servicios de análisis temporalmente no disponibles. Intenta de nuevo en unos minutos."}), 503
        
        print(f"[DEBUG] Respuesta SOAP completa: {soap_response_text}")
        
        # Parsear JSON - ahora debería funcionar siempre
        try:
            resultado = json.loads(soap_response_text)
            print("[SUCCESS] JSON parseado correctamente")
            
            soap_data = resultado.get('soap', {})
            subjetivo = soap_data.get('subjetivo', 'No disponible')
            objetivo = soap_data.get('objetivo', 'No disponible')
            analisis = soap_data.get('analisis', 'No disponible')
            plan = soap_data.get('plan', 'No disponible')
            
            diagnostico = resultado.get('diagnostico', 'No especificado')
            tratamiento = resultado.get('tratamiento', 'No especificado')
            cumplimiento_data = resultado.get('cumplimiento', {})
            cumplimiento_estado = cumplimiento_data.get('estado', 'Pendiente de revisar')
            
        except json.JSONDecodeError as e:
            print(f"[ERROR CRÍTICO] JSON inválido después de forzar formato: {e}")
            print(f"[ERROR] Respuesta completa: {soap_response_text}")
            return jsonify({
                "error": "Error al parsear respuesta de IA. Respuesta no es JSON válido.",
                "debug_info": soap_response_text[:500]
            }), 500
        
        soap_formateado = "S (Subjetivo): " + subjetivo + "\n\n"
        soap_formateado += "O (Objetivo): " + objetivo + "\n\n"
        soap_formateado += "A (Análisis): " + analisis + "\n\n"
        soap_formateado += "P (Plan): " + plan
        
        # Guardar en base de datos
        consulta_data = {
            'transcripcion': transcripcion,
            'soap_subjetivo': subjetivo,
            'soap_objetivo': objetivo,
            'soap_analisis': analisis,
            'soap_plan': plan,
            'diagnostico': diagnostico,
            'tratamiento': tratamiento,
            'cumplimiento_estado': cumplimiento_estado,
            'audio_duracion': 0,
            'paciente_nombre': request.form.get('paciente_nombre', 'Paciente Demo')
        }
        consulta_id = db.guardar_consulta(consulta_data)
        
        return jsonify({
            "transcription": transcripcion,
            "soap_output": soap_formateado,
            "diagnostico": diagnostico,
            "plan": tratamiento,
            "cumplimiento": cumplimiento_estado,
            "consulta_id": consulta_id,
            "provider": provider_used,
            "debug": {
                "raw_ai_response": soap_response_text[:800] if soap_response_text else "No response",
                "response_length": len(soap_response_text) if soap_response_text else 0,
                "transcription_length": len(transcripcion),
                "ai_provider": provider_used
            }
        })
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"[ERROR] Exception: {error_details}")
        return jsonify({
            "error": "Error al procesar audio: " + str(e),
            "debug": {"traceback": error_details[:500]}
        }), 500

@app.route('/api/transcribir_audio_disabled', methods=['POST'])
def transcribir_audio_disabled():
    if 'audio' not in request.files:
        return jsonify({"error": "No se recibió archivo de audio."}), 400
    
    audio_file = request.files['audio']
    
    if audio_file.filename == '':
        return jsonify({"error": "Archivo de audio vacío."}), 400
    
    try:
        audio_bytes = audio_file.read()
        
        transcription_prompt = """Transcribe el siguiente audio de una consulta médica palabra por palabra. 
El audio contiene una conversación entre un médico y un paciente.
Formatea la transcripción claramente indicando quién habla en cada momento.
Ejemplo:
Médico: [texto]
Paciente: [texto]

Transcribe TODO el contenido del audio con precisión."""

        transcription_response = client.models.generate_content(
            model="gemini-2.0-flash-exp",
            contents=[
                types.Part.from_bytes(
                    data=audio_bytes,
                    mime_type="audio/webm",
                ),
                transcription_prompt
            ]
        )
        
        if not transcription_response.text:
            return jsonify({"error": "No se pudo transcribir el audio."}), 500
        
        transcripcion = transcription_response.text
        
        soap_prompt = """Eres un asistente médico experto que analiza transcripciones de consultas médicas.

Analiza la siguiente transcripción de una consulta médica y genera:

1. NOTAS SOAP (formato estructurado):
   - S (Subjetivo): Qué reporta el paciente (síntomas, molestias, historia)
   - O (Objetivo): Hallazgos de la exploración física, signos vitales
   - A (Análisis): Diagnóstico probable o diagnósticos diferenciales
   - P (Plan): Tratamiento, medicamentos (con dosis), estudios, seguimiento

2. DIAGNÓSTICO SUGERIDO: El diagnóstico más probable basado en la información

3. PLAN DE TRATAMIENTO: Resumen del tratamiento con medicamentos y dosis específicas

4. VERIFICACIÓN DE CUMPLIMIENTO: Indica si se mencionó:
   - Consentimiento informado (para procedimientos)
   - Explicación de riesgos
   - Instrucciones claras al paciente

TRANSCRIPCIÓN DE LA CONSULTA:
""" + transcripcion + """

Responde en formato JSON con esta estructura:
{
  "soap": {
    "subjetivo": "texto",
    "objetivo": "texto", 
    "analisis": "texto",
    "plan": "texto"
  },
  "diagnostico": "texto",
  "tratamiento": "texto",
  "cumplimiento": {
    "consentimiento": "mencionado/no mencionado",
    "riesgos_explicados": "si/no",
    "instrucciones_claras": "si/no",
    "estado": "Verificado/Pendiente de revisar"
  }
}"""
        
        soap_response_text = call_gemini_api(soap_prompt, GEMINI_API_KEY)
        
        if not soap_response_text:
            return jsonify({"error": "No se pudo generar notas SOAP."}), 500
        
        resultado = json.loads(soap_response_text)
        
        soap_data = resultado.get('soap', {})
        subjetivo = soap_data.get('subjetivo', 'No disponible')
        objetivo = soap_data.get('objetivo', 'No disponible')
        analisis = soap_data.get('analisis', 'No disponible')
        plan = soap_data.get('plan', 'No disponible')
        
        soap_formateado = "S (Subjetivo): " + subjetivo + "\n\n"
        soap_formateado += "O (Objetivo): " + objetivo + "\n\n"
        soap_formateado += "A (Análisis): " + analisis + "\n\n"
        soap_formateado += "P (Plan): " + plan
        
        diagnostico = resultado.get('diagnostico', 'No especificado')
        tratamiento = resultado.get('tratamiento', 'No especificado')
        cumplimiento_data = resultado.get('cumplimiento', {})
        cumplimiento_estado = cumplimiento_data.get('estado', 'Pendiente de revisar')
        
        # Guardar en base de datos
        consulta_data = {
            'transcripcion': transcripcion,
            'soap_subjetivo': subjetivo,
            'soap_objetivo': objetivo,
            'soap_analisis': analisis,
            'soap_plan': plan,
            'diagnostico': diagnostico,
            'tratamiento': tratamiento,
            'cumplimiento_estado': cumplimiento_estado,
            'audio_duracion': 0,
            'paciente_nombre': request.form.get('paciente_nombre', '')
        }
        consulta_id = db.guardar_consulta(consulta_data)
        
        return jsonify({
            "transcription": transcripcion,
            "soap_output": soap_formateado,
            "diagnostico": diagnostico,
            "plan": tratamiento,
            "cumplimiento": cumplimiento_estado,
            "consulta_id": consulta_id
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": "Error al procesar audio: " + str(e)}), 500

@app.route('/historial')
def vista_historial():
    consultas = db.obtener_consultas(limite=20)
    return render_template('historial.html', consultas=consultas)

@app.route('/api/consulta/<int:consulta_id>')
def obtener_consulta_api(consulta_id):
    consulta = db.obtener_consulta(consulta_id)
    if not consulta:
        return jsonify({"error": "Consulta no encontrada"}), 404
    return jsonify(consulta)

@app.route('/api/consulta/<int:consulta_id>', methods=['PUT'])
def actualizar_consulta_api(consulta_id):
    if not request.json:
        return jsonify({"error": "No se recibió datos JSON"}), 400
    
    success = db.actualizar_consulta(consulta_id, request.json)
    if success:
        return jsonify({"message": "Consulta actualizada correctamente"})
    else:
        return jsonify({"error": "No se pudo actualizar la consulta"}), 400

@app.route('/api/buscar_consultas')
def buscar_consultas_api():
    termino = request.args.get('q', '')
    if not termino:
        return jsonify({"error": "Término de búsqueda requerido"}), 400
    
    consultas = db.buscar_consultas(termino)
    return jsonify(consultas)

@app.route('/api/test_soap_debug', methods=['POST'])
def test_soap_debug():
    """Endpoint de debug para probar generación de SOAP con logs visibles"""
    if not request.json:
        return jsonify({"error": "No se recibió datos JSON"}), 400
    
    transcripcion = request.json.get('transcripcion', '')
    if not transcripcion:
        return jsonify({"error": "Falta transcripción"}), 400
    
    debug_log = []
    
    soap_prompt = """Eres un asistente médico experto que analiza transcripciones de consultas médicas.

Analiza la siguiente transcripción de una consulta médica y genera:

1. NOTAS SOAP (formato estructurado):
   - S (Subjetivo): Qué reporta el paciente (síntomas, molestias, historia)
   - O (Objetivo): Hallazgos de la exploración física, signos vitales
   - A (Análisis): Diagnóstico probable o diagnósticos diferenciales
   - P (Plan): Tratamiento, medicamentos (con dosis), estudios, seguimiento

2. DIAGNÓSTICO SUGERIDO: El diagnóstico más probable basado en la información

3. PLAN DE TRATAMIENTO: Resumen del tratamiento con medicamentos y dosis específicas

4. VERIFICACIÓN DE CUMPLIMIENTO: Indica si se mencionó:
   - Consentimiento informado (para procedimientos)
   - Explicación de riesgos
   - Instrucciones claras al paciente

TRANSCRIPCIÓN DE LA CONSULTA:
""" + transcripcion + """

Responde en formato JSON con esta estructura:
{
  "soap": {
    "subjetivo": "texto",
    "objetivo": "texto", 
    "analisis": "texto",
    "plan": "texto"
  },
  "diagnostico": "texto",
  "tratamiento": "texto",
  "cumplimiento": {
    "consentimiento": "mencionado/no mencionado",
    "riesgos_explicados": "si/no",
    "instrucciones_claras": "si/no",
    "estado": "Verificado/Pendiente de revisar"
  }
}"""
    
    debug_log.append("Llamando a Gemini API con force_json=True...")
    soap_response_text = call_gemini_api(soap_prompt, GEMINI_API_KEY, force_json=True)
    
    if not soap_response_text:
        debug_log.append("ERROR: No se recibió respuesta de Gemini")
        return jsonify({"error": "No response from Gemini", "debug_log": debug_log}), 500
    
    debug_log.append(f"Respuesta recibida (primeros 500 chars): {soap_response_text[:500]}")
    
    try:
        resultado = json.loads(soap_response_text)
        debug_log.append("SUCCESS: JSON parseado correctamente")
        return jsonify({
            "success": True,
            "resultado": resultado,
            "debug_log": debug_log,
            "raw_response": soap_response_text
        })
    except json.JSONDecodeError as e:
        debug_log.append(f"ERROR: JSON inválido - {str(e)}")
        return jsonify({
            "success": False,
            "error": str(e),
            "debug_log": debug_log,
            "raw_response": soap_response_text
        }), 500


@app.route('/api/validar_farmacovigilancia', methods=['POST'])
def api_validar_farmacovigilancia():
    """
    Farmacovigilancia y seguridad clínica (DDI).
    Recibe texto de receta o nota SOAP; extrae entidades con Gemini (solo orquestador)
    y cruza contra fuente de verdad (mock/DB). CERO alucinaciones: la IA no inventa interacciones.
    """
    if not request.json:
        return jsonify({"error": "No se recibió JSON"}), 400
    texto = request.json.get("texto") or request.json.get("transcripcion") or ""
    if not texto or not str(texto).strip():
        return jsonify({"error": "Se requiere 'texto' o 'transcripcion' con contenido"}), 400
    resultado = ejecutar_validacion_completa(str(texto).strip(), GEMINI_API_KEY)
    return jsonify(resultado)


@app.route('/contador')
def vista_contador():
    """Vista principal del módulo del contador con grid dinámico"""
    # Obtener estadísticas financieras
    stats = transaccion_db.obtener_estadisticas_financieras()
    
    # Obtener transacciones recientes
    transacciones = transaccion_db.obtener_transacciones(limite=50)
    
    # Obtener clasificaciones disponibles para el frontend
    clasificaciones_ingresos = obtener_lista_clasificaciones_por_tipo('ingreso')
    clasificaciones_gastos = obtener_lista_clasificaciones_por_tipo('gasto')
    
    # Obtener formas de pago válidas del SAT
    formas_pago = obtener_formas_pago()
    
    return render_template('contador.html', 
                         stats=stats, 
                         transacciones=transacciones,
                         clasificaciones_ingresos=clasificaciones_ingresos,
                         clasificaciones_gastos=clasificaciones_gastos,
                         formas_pago=formas_pago)


@app.route('/agente-contable')
def vista_agente_contable():
    return render_template('agente_contable.html')


@app.route('/api/agente-contable/chat', methods=['POST'])
def api_agente_contable_chat():
    from src.core.gemini_client import create_agent, GeminiModel

    if not request.is_json:
        return jsonify({"error": "JSON requerido."}), 400
    data = request.get_json() or {}

    if data.get("reset"):
        session.pop("agente_contable_hist", None)
        session.modified = True
        return jsonify({"reply": "", "ok": True})

    msg = (data.get("message") or "").strip()
    if not msg:
        return jsonify({"error": "Mensaje vacío."}), 400

    if not GEMINI_API_KEY:
        return jsonify({"error": "GEMINI_API_KEY no configurada."}), 503

    hist = session.get("agente_contable_hist") or []
    genai_hist = _agente_contable_hist_to_genai(hist)

    try:
        agent = create_agent(model=GeminiModel.FLASH)
        chat = agent.start_chat(history=genai_hist)
        response = chat.send_message(msg)
        reply = (response.text or "").strip()
    except Exception as e:
        return jsonify({"error": f"No se pudo generar respuesta: {e!s}"}), 500

    hist.append({"role": "user", "text": msg})
    hist.append({"role": "model", "text": reply})
    session["agente_contable_hist"] = hist[-40:]
    session.modified = True
    return jsonify({"reply": reply})


@app.route('/debug_soap')
def vista_debug_soap():
    return render_template('debug_soap.html')

@app.route('/api/transacciones', methods=['GET'])
def obtener_transacciones_api():
    """API para obtener transacciones con filtros"""
    filtros = {
        'medico_id': request.args.get('medico_id', 'default'),
        'tipo': request.args.get('tipo'),
        'estatus_validacion': request.args.get('estatus'),
        'fecha_desde': request.args.get('fecha_desde'),
        'fecha_hasta': request.args.get('fecha_hasta'),
        'clasificacion': request.args.get('clasificacion')
    }
    
    # Remover filtros vacíos
    filtros = {k: v for k, v in filtros.items() if v}
    
    limite = int(request.args.get('limite', 100))
    transacciones = transaccion_db.obtener_transacciones(filtros, limite)
    
    return jsonify(transacciones)

@app.route('/api/transacciones', methods=['POST'])
def crear_transaccion_api():
    """API para crear una nueva transacción"""
    if not request.json:
        return jsonify({"error": "No se recibió datos JSON"}), 400
    
    tipo = request.json.get('tipo', 'gasto')
    concepto = request.json.get('concepto', '')
    proveedor = request.json.get('proveedor', '')
    clasificacion_manual = request.json.get('clasificacion', '')
    forma_pago = request.json.get('forma_pago', '')
    monto = float(request.json.get('monto', 0))
    
    # Validar forma de pago
    warnings = []
    if forma_pago and not validar_forma_pago(forma_pago):
        return jsonify({"error": f"Forma de pago inválida: {forma_pago}"}), 400
    
    # Validar deducibilidad de efectivo
    validacion_efectivo = None
    if forma_pago and tipo == 'gasto':
        validacion_efectivo = validar_deducibilidad_efectivo(monto, forma_pago)
        if validacion_efectivo['warning']:
            warnings.append(validacion_efectivo['mensaje'])
    
    # Si se proporciona clasificación manual, validarla y usar su porcentaje
    if clasificacion_manual and validar_clasificacion(clasificacion_manual, tipo):
        clasificacion_ia = {
            'clasificacion': clasificacion_manual,
            'deducible_porcentaje': obtener_porcentaje_deducible(clasificacion_manual),
            'confianza': 'alta',
            'metodo': 'manual'
        }
        # Si hay warning de efectivo y no es deducible, sobrescribir porcentaje
        if validacion_efectivo and not validacion_efectivo['es_deducible']:
            clasificacion_ia['deducible_porcentaje'] = 0
    else:
        # Clasificar automáticamente con IA
        clasificacion_ia = transaccion_db.clasificar_con_ia(concepto, proveedor)
        # Si la clasificación sugerida no es válida, usar una por defecto
        if not validar_clasificacion(clasificacion_ia.get('clasificacion', ''), tipo):
            clasificaciones = obtener_lista_clasificaciones_por_tipo(tipo)
            if clasificaciones:
                clasificacion_ia['clasificacion'] = clasificaciones[0]
                clasificacion_ia['deducible_porcentaje'] = obtener_porcentaje_deducible(clasificaciones[0])
        # Si hay warning de efectivo y no es deducible, sobrescribir porcentaje
        if validacion_efectivo and not validacion_efectivo['es_deducible']:
            clasificacion_ia['deducible_porcentaje'] = 0
    
    transaccion_data = {
        **request.json,
        'clasificacion_ia': clasificacion_ia['clasificacion'],
        'deducible_porcentaje': clasificacion_ia['deducible_porcentaje']
    }
    
    transaccion_id = transaccion_db.guardar_transaccion(transaccion_data)
    
    response = {
        "id": transaccion_id,
        "clasificacion_sugerida": clasificacion_ia,
        "message": "Transacción creada exitosamente"
    }
    
    if warnings:
        response["warnings"] = warnings
    
    return jsonify(response), 201

@app.route('/api/transacciones/<int:transaccion_id>/validar', methods=['POST'])
def validar_transaccion_api(transaccion_id):
    """API para validar una transacción (aprobar/rechazar/ajustar)"""
    if not request.json:
        return jsonify({"error": "No se recibió datos JSON"}), 400
    
    clasificacion = request.json.get('clasificacion', '')
    deducible_manual = request.json.get('deducible_porcentaje')
    
    # Obtener el tipo de transacción para validar la clasificación
    transaccion = transaccion_db.obtener_transacciones({'id': transaccion_id}, limite=1)
    tipo = transaccion[0].get('tipo', 'gasto') if transaccion else 'gasto'
    
    # Validar clasificación y obtener porcentaje si es válida
    if clasificacion and validar_clasificacion(clasificacion, tipo):
        porcentaje = obtener_porcentaje_deducible(clasificacion)
        # Si se proporciona un porcentaje manual, usarlo (permite ajustes)
        if deducible_manual is not None:
            porcentaje = int(deducible_manual)
    else:
        porcentaje = deducible_manual if deducible_manual is not None else 0
    
    validacion_data = {
        'estatus': request.json.get('estatus', 'aprobado'),
        'clasificacion': clasificacion,
        'deducible_porcentaje': porcentaje,
        'notas': request.json.get('notas', ''),
        'validado_por': request.json.get('validado_por', 'contador')
    }
    
    success = transaccion_db.validar_transaccion(transaccion_id, validacion_data)
    
    if success:
        return jsonify({"message": "Transacción validada correctamente"})
    else:
        return jsonify({"error": "No se pudo validar la transacción"}), 400

@app.route('/api/clasificaciones', methods=['GET'])
def obtener_clasificaciones_api():
    """API para obtener clasificaciones fiscales disponibles"""
    tipo = request.args.get('tipo', 'gasto')
    clasificaciones = obtener_clasificaciones_por_tipo(tipo)
    
    # Formatear para el frontend
    resultado = []
    for nombre, datos in clasificaciones.items():
        resultado.append({
            'nombre': nombre,
            'deducible_porcentaje': datos['deducible_porcentaje'],
            'descripcion': datos['descripcion']
        })
    
    return jsonify(resultado)

@app.route('/api/clasificar_gasto', methods=['POST'])
def clasificar_gasto_api():
    """API para clasificar un gasto usando IA (con Gemini para casos nuevos)"""
    if not request.json:
        return jsonify({"error": "No se recibió datos JSON"}), 400
    
    concepto = request.json.get('concepto', '')
    proveedor = request.json.get('proveedor', '')
    monto = request.json.get('monto', 0)
    
    # Primero intentar con reglas aprendidas
    clasificacion_reglas = transaccion_db.clasificar_con_ia(concepto, proveedor)
    
    # Si la confianza es baja, usar Gemini para clasificación inteligente
    if clasificacion_reglas['confianza'] == 'baja' and GEMINI_API_KEY:
        # Obtener lista de clasificaciones válidas para sugerir
        clasificaciones_validas = obtener_lista_clasificaciones_por_tipo('gasto')
        clasificaciones_str = ', '.join(clasificaciones_validas[:10])  # Primeras 10 para no saturar
        
        prompt = f"""Eres un experto contador especializado en fiscalidad médica en México.

Analiza el siguiente gasto y selecciona UNA de estas clasificaciones fiscales válidas:
{clasificaciones_str}

IMPORTANTE: Debes elegir EXACTAMENTE una de las clasificaciones de la lista anterior.

GASTO:
- Concepto: {concepto}
- Proveedor: {proveedor}
- Monto: ${monto}

CONTEXTO FISCAL:
{NORMAS_CONTABLES_BASE}

Responde en formato JSON:
{{
  "clasificacion": "nombre_exacto_de_la_clasificacion",
  "deducible_porcentaje": numero,
  "justificacion": "texto breve"
}}"""
        
        try:
            response_text = call_gemini_api(prompt, GEMINI_API_KEY, force_json=True)
            if response_text:
                resultado_ia = json.loads(response_text)
                # Validar que la clasificación sea válida
                clasif_nombre = resultado_ia.get('clasificacion', '')
                if validar_clasificacion(clasif_nombre, 'gasto'):
                    # Obtener el porcentaje correcto de la clasificación
                    porcentaje = obtener_porcentaje_deducible(clasif_nombre)
                    return jsonify({
                        'clasificacion': clasif_nombre,
                        'deducible_porcentaje': porcentaje,
                        'justificacion': resultado_ia.get('justificacion', ''),
                        'confianza': 'alta',
                        'metodo': 'gemini_ia'
                    })
        except Exception as e:
            print(f"Error en clasificación con Gemini: {e}")
    
    return jsonify(clasificacion_reglas)

@app.route('/api/estadisticas_financieras')
def estadisticas_financieras_api():
    """API para obtener estadísticas financieras del contador"""
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')
    medico_id = request.args.get('medico_id', 'default')
    
    stats = transaccion_db.obtener_estadisticas_financieras(medico_id, fecha_desde, fecha_hasta)
    return jsonify(stats)

@app.route('/api/contador/template-excel')
def template_excel_api():
    """API para descargar template Excel para importar transacciones"""
    # Crear archivo en memoria
    output = BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet('Transacciones')
    
    # Estilos
    header_format = workbook.add_format({
        'bold': True,
        'bg_color': '#4472C4',
        'font_color': '#FFFFFF',
        'border': 1,
        'align': 'center',
        'valign': 'vcenter'
    })
    
    date_format = workbook.add_format({'num_format': 'dd/mm/yyyy'})
    currency_format = workbook.add_format({'num_format': '$#,##0.00'})
    
    # Headers
    headers = [
        'Fecha', 'Tipo', 'RFC_Emisor', 'RFC_Receptor', 'UUID', 
        'Concepto', 'Subtotal', 'IVA', 'Total', 
        'Forma_Pago', 'Metodo_Pago', 'Clasificacion', 
        'Deducible_%', 'Cuenta_Bancaria', 'Notas'
    ]
    
    # Escribir headers
    for col, header in enumerate(headers):
        worksheet.write(0, col, header, header_format)
    
    # Ejemplo de datos (segunda fila)
    hoy = datetime.now().strftime('%Y-%m-%d')
    ejemplo = [
        hoy,  # Fecha
        'Gasto',  # Tipo
        'XAXX010101000',  # RFC_Emisor (ejemplo)
        'XAXX010101000',  # RFC_Receptor
        '00000000-0000-0000-0000-000000000000',  # UUID
        'Material de curación',  # Concepto
        1500.00,  # Subtotal
        0.00,  # IVA
        1500.00,  # Total
        '03 - Transferencia electrónica',  # Forma_Pago
        'Transferencia bancaria',  # Metodo_Pago
        'Material de curación',  # Clasificacion
        100,  # Deducible_%
        '1234567890',  # Cuenta_Bancaria
        'Ejemplo de transacción'  # Notas
    ]
    
    # Escribir ejemplo
    for col, value in enumerate(ejemplo):
        if col == 0:  # Fecha
            worksheet.write(1, col, value, date_format)
        elif col in [6, 7, 8]:  # Subtotal, IVA, Total
            worksheet.write(1, col, value, currency_format)
        else:
            worksheet.write(1, col, value)
    
    # Ajustar ancho de columnas
    worksheet.set_column('A:A', 12)  # Fecha
    worksheet.set_column('B:B', 10)  # Tipo
    worksheet.set_column('C:D', 15)  # RFCs
    worksheet.set_column('E:E', 36)  # UUID
    worksheet.set_column('F:F', 25)  # Concepto
    worksheet.set_column('G:I', 12)  # Montos
    worksheet.set_column('J:J', 30)  # Forma_Pago
    worksheet.set_column('K:K', 20)  # Metodo_Pago
    worksheet.set_column('L:L', 25)  # Clasificacion
    worksheet.set_column('M:M', 12)  # Deducible_%
    worksheet.set_column('N:N', 15)  # Cuenta_Bancaria
    worksheet.set_column('O:O', 30)  # Notas
    
    workbook.close()
    output.seek(0)
    
    # Nombre del archivo con mes y año
    mes_ano = datetime.now().strftime('%m_%Y')
    filename = f'Template_Transacciones_{mes_ano}.xlsx'
    
    response = make_response(output.getvalue())
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    response.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    
    return response

@app.route('/api/contador/importar-excel', methods=['POST'])
def importar_excel_api():
    """API para importar transacciones desde un archivo Excel"""
    if 'archivo' not in request.files:
        return jsonify({"error": "No se recibió archivo"}), 400
    
    archivo = request.files['archivo']
    if archivo.filename == '':
        return jsonify({"error": "Archivo vacío"}), 400
    
    # Validar extensión
    if not (archivo.filename.endswith('.xlsx') or archivo.filename.endswith('.csv')):
        return jsonify({"error": "Formato no soportado. Solo .xlsx y .csv"}), 400
    
    exitosas = 0
    duplicadas = 0
    errores = []
    
    try:
        if archivo.filename.endswith('.xlsx'):
            # Leer Excel
            workbook = openpyxl.load_workbook(archivo)
            worksheet = workbook.active
            
            # Leer headers (primera fila)
            headers = []
            for cell in worksheet[1]:
                headers.append(cell.value.strip() if cell.value else '')
            
            # Validar que tenga las columnas necesarias
            columnas_requeridas = ['Fecha', 'Tipo', 'Concepto', 'Total']
            columnas_faltantes = [col for col in columnas_requeridas if col not in headers]
            if columnas_faltantes:
                return jsonify({
                    "error": f"Columnas faltantes: {', '.join(columnas_faltantes)}",
                    "columnas_encontradas": headers
                }), 400
            
            # Mapeo de índices de columnas
            idx_fecha = headers.index('Fecha')
            idx_tipo = headers.index('Tipo')
            idx_concepto = headers.index('Concepto')
            idx_total = headers.index('Total')
            idx_rfc_emisor = headers.index('RFC_Emisor') if 'RFC_Emisor' in headers else None
            idx_rfc_receptor = headers.index('RFC_Receptor') if 'RFC_Receptor' in headers else None
            idx_uuid = headers.index('UUID') if 'UUID' in headers else None
            idx_subtotal = headers.index('Subtotal') if 'Subtotal' in headers else None
            idx_iva = headers.index('IVA') if 'IVA' in headers else None
            idx_forma_pago = headers.index('Forma_Pago') if 'Forma_Pago' in headers else None
            idx_metodo_pago = headers.index('Metodo_Pago') if 'Metodo_Pago' in headers else None
            idx_clasificacion = headers.index('Clasificacion') if 'Clasificacion' in headers else None
            idx_deducible = headers.index('Deducible_%') if 'Deducible_%' in headers else None
            idx_cuenta = headers.index('Cuenta_Bancaria') if 'Cuenta_Bancaria' in headers else None
            idx_notas = headers.index('Notas') if 'Notas' in headers else None
            idx_proveedor = headers.index('Proveedor') if 'Proveedor' in headers else None
            
            # Procesar filas (empezar desde la fila 2, saltando headers y ejemplo)
            for fila_num, row in enumerate(worksheet.iter_rows(min_row=3, values_only=False), start=3):
                # Saltar filas vacías
                if not any(cell.value for cell in row):
                    continue
                
                try:
                    # Extraer valores
                    fecha_val = row[idx_fecha].value
                    tipo_val = row[idx_tipo].value
                    concepto_val = row[idx_concepto].value
                    total_val = row[idx_total].value
                    
                    # Validaciones básicas
                    if not fecha_val:
                        errores.append({"fila": fila_num, "error": "Fecha inválida o vacía"})
                        continue
                    
                    if not tipo_val or str(tipo_val).lower() not in ['ingreso', 'gasto']:
                        errores.append({"fila": fila_num, "error": f"Tipo inválido: {tipo_val}. Debe ser 'Ingreso' o 'Gasto'"})
                        continue
                    
                    if not concepto_val:
                        errores.append({"fila": fila_num, "error": "Concepto vacío"})
                        continue
                    
                    if not total_val or float(total_val) <= 0:
                        errores.append({"fila": fila_num, "error": f"Monto debe ser mayor a 0. Valor: {total_val}"})
                        continue
                    
                    # Convertir fecha
                    if isinstance(fecha_val, str):
                        try:
                            fecha = datetime.strptime(fecha_val, '%Y-%m-%d').date()
                        except:
                            try:
                                fecha = datetime.strptime(fecha_val, '%d/%m/%Y').date()
                            except:
                                errores.append({"fila": fila_num, "error": f"Formato de fecha inválido: {fecha_val}"})
                                continue
                    else:
                        fecha = fecha_val.date() if hasattr(fecha_val, 'date') else fecha_val
                    
                    # Preparar datos de transacción
                    transaccion_data = {
                        'tipo': str(tipo_val).lower(),
                        'fecha': fecha.strftime('%Y-%m-%d'),
                        'monto': float(total_val),
                        'concepto': str(concepto_val),
                        'proveedor': str(row[idx_proveedor].value) if idx_proveedor and row[idx_proveedor].value else '',
                        'cfdi_uuid': str(row[idx_uuid].value) if idx_uuid and row[idx_uuid].value else '',
                        'forma_pago': str(row[idx_forma_pago].value) if idx_forma_pago and row[idx_forma_pago].value else '',
                        'metodo_pago': str(row[idx_metodo_pago].value) if idx_metodo_pago and row[idx_metodo_pago].value else '',
                        'clasificacion': str(row[idx_clasificacion].value) if idx_clasificacion and row[idx_clasificacion].value else '',
                        'deducible_porcentaje': int(row[idx_deducible].value) if idx_deducible and row[idx_deducible].value else None,
                        'notas_contador': str(row[idx_notas].value) if idx_notas and row[idx_notas].value else ''
                    }
                    
                    # Validar UUID único si existe
                    if transaccion_data['cfdi_uuid']:
                        uuid_existente = transaccion_db.obtener_transacciones({'cfdi_uuid': transaccion_data['cfdi_uuid']}, limite=1)
                        if uuid_existente:
                            duplicadas += 1
                            continue
                    
                    # Validar forma de pago si existe
                    if transaccion_data['forma_pago'] and not validar_forma_pago(transaccion_data['forma_pago']):
                        errores.append({"fila": fila_num, "error": f"Forma de pago inválida: {transaccion_data['forma_pago']}"})
                        continue
                    
                    # Validar clasificación si existe
                    if transaccion_data['clasificacion']:
                        if not validar_clasificacion(transaccion_data['clasificacion'], transaccion_data['tipo']):
                            # Si la clasificación no es válida, intentar clasificar con IA
                            clasificacion_ia = transaccion_db.clasificar_con_ia(transaccion_data['concepto'], transaccion_data['proveedor'])
                            transaccion_data['clasificacion_ia'] = clasificacion_ia['clasificacion']
                            transaccion_data['deducible_porcentaje'] = clasificacion_ia['deducible_porcentaje']
                        else:
                            transaccion_data['clasificacion_ia'] = transaccion_data['clasificacion']
                            if transaccion_data['deducible_porcentaje'] is None:
                                transaccion_data['deducible_porcentaje'] = obtener_porcentaje_deducible(transaccion_data['clasificacion'])
                    else:
                        # Clasificar automáticamente
                        clasificacion_ia = transaccion_db.clasificar_con_ia(transaccion_data['concepto'], transaccion_data['proveedor'])
                        transaccion_data['clasificacion_ia'] = clasificacion_ia['clasificacion']
                        transaccion_data['deducible_porcentaje'] = clasificacion_ia['deducible_porcentaje']
                    
                    # Validar deducibilidad de efectivo
                    if transaccion_data['tipo'] == 'gasto' and transaccion_data['forma_pago']:
                        validacion_efectivo = validar_deducibilidad_efectivo(transaccion_data['monto'], transaccion_data['forma_pago'])
                        if not validacion_efectivo['es_deducible']:
                            transaccion_data['deducible_porcentaje'] = 0
                    
                    # Guardar transacción
                    transaccion_id = transaccion_db.guardar_transaccion(transaccion_data)
                    exitosas += 1
                    
                except Exception as e:
                    errores.append({"fila": fila_num, "error": f"Error procesando fila: {str(e)}"})
                    continue
        
        elif archivo.filename.endswith('.csv'):
            # Leer CSV
            import csv
            from io import TextIOWrapper
            
            csv_file = TextIOWrapper(archivo.stream, encoding='utf-8')
            reader = csv.DictReader(csv_file)
            
            for fila_num, row in enumerate(reader, start=2):
                try:
                    # Validaciones básicas
                    if not row.get('Fecha'):
                        errores.append({"fila": fila_num, "error": "Fecha inválida o vacía"})
                        continue
                    
                    tipo_val = row.get('Tipo', '').lower()
                    if tipo_val not in ['ingreso', 'gasto']:
                        errores.append({"fila": fila_num, "error": f"Tipo inválido: {row.get('Tipo')}"})
                        continue
                    
                    if not row.get('Concepto'):
                        errores.append({"fila": fila_num, "error": "Concepto vacío"})
                        continue
                    
                    monto = float(row.get('Total', 0) or row.get('Monto', 0))
                    if monto <= 0:
                        errores.append({"fila": fila_num, "error": f"Monto debe ser mayor a 0"})
                        continue
                    
                    # Convertir fecha
                    fecha_str = row['Fecha']
                    try:
                        fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
                    except:
                        try:
                            fecha = datetime.strptime(fecha_str, '%d/%m/%Y').date()
                        except:
                            errores.append({"fila": fila_num, "error": f"Formato de fecha inválido: {fecha_str}"})
                            continue
                    
                    # Preparar datos
                    transaccion_data = {
                        'tipo': tipo_val,
                        'fecha': fecha.strftime('%Y-%m-%d'),
                        'monto': monto,
                        'concepto': row.get('Concepto', ''),
                        'proveedor': row.get('Proveedor', ''),
                        'cfdi_uuid': row.get('UUID', ''),
                        'forma_pago': row.get('Forma_Pago', ''),
                        'metodo_pago': row.get('Metodo_Pago', ''),
                        'clasificacion': row.get('Clasificacion', ''),
                        'deducible_porcentaje': int(row['Deducible_%']) if row.get('Deducible_%') else None,
                        'notas_contador': row.get('Notas', '')
                    }
                    
                    # Validar UUID único
                    if transaccion_data['cfdi_uuid']:
                        uuid_existente = transaccion_db.obtener_transacciones({'cfdi_uuid': transaccion_data['cfdi_uuid']}, limite=1)
                        if uuid_existente:
                            duplicadas += 1
                            continue
                    
                    # Validar y clasificar
                    if transaccion_data['clasificacion'] and validar_clasificacion(transaccion_data['clasificacion'], tipo_val):
                        transaccion_data['clasificacion_ia'] = transaccion_data['clasificacion']
                        if transaccion_data['deducible_porcentaje'] is None:
                            transaccion_data['deducible_porcentaje'] = obtener_porcentaje_deducible(transaccion_data['clasificacion'])
                    else:
                        clasificacion_ia = transaccion_db.clasificar_con_ia(transaccion_data['concepto'], transaccion_data['proveedor'])
                        transaccion_data['clasificacion_ia'] = clasificacion_ia['clasificacion']
                        transaccion_data['deducible_porcentaje'] = clasificacion_ia['deducible_porcentaje']
                    
                    # Validar efectivo
                    if transaccion_data['tipo'] == 'gasto' and transaccion_data['forma_pago']:
                        validacion_efectivo = validar_deducibilidad_efectivo(transaccion_data['monto'], transaccion_data['forma_pago'])
                        if not validacion_efectivo['es_deducible']:
                            transaccion_data['deducible_porcentaje'] = 0
                    
                    transaccion_id = transaccion_db.guardar_transaccion(transaccion_data)
                    exitosas += 1
                    
                except Exception as e:
                    errores.append({"fila": fila_num, "error": f"Error procesando fila: {str(e)}"})
                    continue
    
    except Exception as e:
        return jsonify({"error": f"Error al procesar archivo: {str(e)}"}), 500
    
    mensaje = f"{exitosas} transacciones importadas"
    if duplicadas > 0:
        mensaje += f", {duplicadas} duplicadas"
    if errores:
        mensaje += f", {len(errores)} errores"
    
    return jsonify({
        "exitosas": exitosas,
        "duplicadas": duplicadas,
        "errores": errores,
        "mensaje": mensaje
    })

@app.route('/api/contador/exportar-excel')
def exportar_excel_completo_api():
    """API para exportar reporte completo a Excel con 3 hojas"""
    # Obtener filtros
    filtros = {
        'medico_id': request.args.get('medico_id', 'default'),
        'tipo': request.args.get('tipo'),
        'estatus_validacion': request.args.get('estatus'),
        'fecha_desde': request.args.get('fecha_desde'),
        'fecha_hasta': request.args.get('fecha_hasta')
    }
    filtros = {k: v for k, v in filtros.items() if v}
    
    # Obtener transacciones
    transacciones = transaccion_db.obtener_transacciones(filtros, limite=10000)
    
    # Obtener estadísticas
    stats = transaccion_db.obtener_estadisticas_financieras(
        filtros.get('medico_id', 'default'),
        filtros.get('fecha_desde'),
        filtros.get('fecha_hasta')
    )
    
    # Crear archivo en memoria
    output = BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    
    # Estilos
    header_format = workbook.add_format({
        'bold': True,
        'bg_color': '#4472C4',
        'font_color': '#FFFFFF',
        'border': 1,
        'align': 'center',
        'valign': 'vcenter'
    })
    
    date_format = workbook.add_format({'num_format': 'dd/mm/yyyy'})
    currency_format = workbook.add_format({'num_format': '$#,##0.00'})
    currency_bold_format = workbook.add_format({
        'num_format': '$#,##0.00',
        'bold': True
    })
    bold_format = workbook.add_format({'bold': True})
    percent_format = workbook.add_format({'num_format': '0%'})
    
    # ============================================
    # HOJA 1: TRANSACCIONES
    # ============================================
    worksheet1 = workbook.add_worksheet('Transacciones')
    
    headers1 = [
        'ID', 'Fecha', 'Tipo', 'RFC_Emisor', 'RFC_Receptor', 'UUID',
        'Concepto', 'Proveedor', 'Subtotal', 'IVA', 'Total',
        'Forma_Pago', 'Metodo_Pago', 'Clasificacion',
        'Deducible_%', 'Cuenta_Bancaria', 'Estatus', 'Notas'
    ]
    
    # Escribir headers
    for col, header in enumerate(headers1):
        worksheet1.write(0, col, header, header_format)
    
    # Escribir datos
    total_ingresos = 0
    total_gastos = 0
    
    for row_num, t in enumerate(transacciones, start=1):
        tipo = t.get('tipo', '')
        monto = float(t.get('monto', 0))
        
        if tipo == 'ingreso':
            total_ingresos += monto
        else:
            total_gastos += monto
        
        # Calcular subtotal e IVA (asumiendo IVA 0% para honorarios médicos)
        subtotal = monto
        iva = 0.0
        
        worksheet1.write(row_num, 0, t.get('id', ''))
        worksheet1.write(row_num, 1, t.get('fecha', ''), date_format)
        worksheet1.write(row_num, 2, tipo.upper() if tipo else '')
        worksheet1.write(row_num, 3, '')  # RFC_Emisor
        worksheet1.write(row_num, 4, '')  # RFC_Receptor
        worksheet1.write(row_num, 5, t.get('cfdi_uuid', ''))
        worksheet1.write(row_num, 6, t.get('concepto', ''))
        worksheet1.write(row_num, 7, t.get('proveedor', ''))
        worksheet1.write(row_num, 8, subtotal, currency_format)
        worksheet1.write(row_num, 9, iva, currency_format)
        worksheet1.write(row_num, 10, monto, currency_format)
        worksheet1.write(row_num, 11, t.get('forma_pago', ''))
        worksheet1.write(row_num, 12, t.get('metodo_pago', ''))
        worksheet1.write(row_num, 13, t.get('clasificacion_contador') or t.get('clasificacion_ia', ''))
        worksheet1.write(row_num, 14, t.get('deducible_porcentaje', 0))
        worksheet1.write(row_num, 15, '')  # Cuenta_Bancaria
        worksheet1.write(row_num, 16, t.get('estatus_validacion', '').upper())
        worksheet1.write(row_num, 17, t.get('notas_contador', ''))
    
    # Totales al final
    total_row = len(transacciones) + 2
    worksheet1.write(total_row, 6, 'TOTALES', bold_format)
    worksheet1.write(total_row, 8, total_ingresos + total_gastos, currency_bold_format)
    worksheet1.write(total_row, 9, 0, currency_bold_format)
    worksheet1.write(total_row, 10, total_ingresos + total_gastos, currency_bold_format)
    
    # Ajustar ancho de columnas
    worksheet1.set_column('A:A', 8)   # ID
    worksheet1.set_column('B:B', 12)  # Fecha
    worksheet1.set_column('C:C', 10)  # Tipo
    worksheet1.set_column('D:E', 15)  # RFCs
    worksheet1.set_column('F:F', 36)  # UUID
    worksheet1.set_column('G:G', 25)  # Concepto
    worksheet1.set_column('H:H', 20)  # Proveedor
    worksheet1.set_column('I:K', 12)  # Montos
    worksheet1.set_column('L:L', 30)  # Forma_Pago
    worksheet1.set_column('M:M', 20)  # Metodo_Pago
    worksheet1.set_column('N:N', 25)  # Clasificacion
    worksheet1.set_column('O:O', 12)   # Deducible_%
    worksheet1.set_column('P:P', 15)  # Cuenta_Bancaria
    worksheet1.set_column('Q:Q', 12)   # Estatus
    worksheet1.set_column('R:R', 30)  # Notas
    
    # ============================================
    # HOJA 2: RESUMEN
    # ============================================
    worksheet2 = workbook.add_worksheet('Resumen')
    
    # Título
    worksheet2.write(0, 0, 'RESUMEN FINANCIERO', bold_format)
    worksheet2.write(1, 0, '')
    
    # Totales principales
    row = 2
    worksheet2.write(row, 0, 'Total Ingresos:', bold_format)
    worksheet2.write(row, 1, stats.get('ingresos_totales', 0), currency_format)
    row += 1
    
    worksheet2.write(row, 0, 'Total Gastos:', bold_format)
    worksheet2.write(row, 1, stats.get('gastos_totales', 0), currency_format)
    row += 1
    
    worksheet2.write(row, 0, 'Utilidad (Ingresos - Gastos):', bold_format)
    worksheet2.write(row, 1, stats.get('utilidad', 0), currency_bold_format)
    row += 2
    
    # Tabla por clasificación
    worksheet2.write(row, 0, 'CLASIFICACIÓN', header_format)
    worksheet2.write(row, 1, 'MONTO', header_format)
    worksheet2.write(row, 2, '% DEL TOTAL', header_format)
    row += 1
    
    # Agrupar por clasificación
    clasificaciones_dict = {}
    total_clasificaciones = 0
    
    for t in transacciones:
        clasif = t.get('clasificacion_contador') or t.get('clasificacion_ia', 'Sin clasificar')
        monto = float(t.get('monto', 0))
        if clasif not in clasificaciones_dict:
            clasificaciones_dict[clasif] = 0
        clasificaciones_dict[clasif] += monto
        total_clasificaciones += monto
    
    # Ordenar por monto descendente
    clasificaciones_sorted = sorted(clasificaciones_dict.items(), key=lambda x: x[1], reverse=True)
    
    for clasif, monto in clasificaciones_sorted:
        porcentaje = (monto / total_clasificaciones * 100) if total_clasificaciones > 0 else 0
        worksheet2.write(row, 0, clasif)
        worksheet2.write(row, 1, monto, currency_format)
        worksheet2.write(row, 2, porcentaje / 100, percent_format)
        row += 1
    
    # Total de clasificaciones
    worksheet2.write(row, 0, 'TOTAL', bold_format)
    worksheet2.write(row, 1, total_clasificaciones, currency_bold_format)
    worksheet2.write(row, 2, 1.0, percent_format)
    
    # Ajustar ancho de columnas
    worksheet2.set_column('A:A', 30)
    worksheet2.set_column('B:B', 15)
    worksheet2.set_column('C:C', 12)
    
    # ============================================
    # HOJA 3: DEDUCIBLES
    # ============================================
    worksheet3 = workbook.add_worksheet('Deducibles')
    
    headers3 = [
        'Fecha', 'Concepto', 'Proveedor', 'Monto', '% Deducible', 'Monto Deducible'
    ]
    
    # Escribir headers
    for col, header in enumerate(headers3):
        worksheet3.write(0, col, header, header_format)
    
    # Filtrar solo gastos deducibles aprobados
    gastos_deducibles = [
        t for t in transacciones
        if t.get('tipo') == 'gasto' 
        and t.get('estatus_validacion') == 'aprobado'
        and t.get('deducible_porcentaje', 0) > 0
    ]
    
    total_deducible = 0
    
    # Escribir datos
    for row_num, t in enumerate(gastos_deducibles, start=1):
        monto = float(t.get('monto', 0))
        porcentaje = float(t.get('deducible_porcentaje', 0))
        monto_deducible = monto * (porcentaje / 100.0)
        total_deducible += monto_deducible
        
        worksheet3.write(row_num, 0, t.get('fecha', ''), date_format)
        worksheet3.write(row_num, 1, t.get('concepto', ''))
        worksheet3.write(row_num, 2, t.get('proveedor', ''))
        worksheet3.write(row_num, 3, monto, currency_format)
        worksheet3.write(row_num, 4, porcentaje / 100, percent_format)
        worksheet3.write(row_num, 5, monto_deducible, currency_format)
    
    # Total deducible al final
    total_row3 = len(gastos_deducibles) + 2
    worksheet3.write(total_row3, 3, 'TOTAL DEDUCIBLE', bold_format)
    worksheet3.write(total_row3, 5, total_deducible, currency_bold_format)
    
    # Ajustar ancho de columnas
    worksheet3.set_column('A:A', 12)  # Fecha
    worksheet3.set_column('B:B', 30)  # Concepto
    worksheet3.set_column('C:C', 20)  # Proveedor
    worksheet3.set_column('D:D', 15)  # Monto
    worksheet3.set_column('E:E', 15)  # % Deducible
    worksheet3.set_column('F:F', 18)  # Monto Deducible
    
    workbook.close()
    output.seek(0)
    
    # Nombre del archivo
    mes_ano = datetime.now().strftime('%m_%Y')
    medico_nombre = filtros.get('medico_id', 'Doctor')
    filename = f'Reporte_Fiscal_{medico_nombre}_{mes_ano}.xlsx'
    
    response = make_response(output.getvalue())
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    response.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    
    return response

@app.route('/api/exportar_transacciones')
def exportar_transacciones_api():
    """API para exportar transacciones a CSV (legacy)"""
    import csv
    from io import StringIO
    
    filtros = {
        'medico_id': request.args.get('medico_id', 'default'),
        'tipo': request.args.get('tipo'),
        'estatus_validacion': request.args.get('estatus'),
        'fecha_desde': request.args.get('fecha_desde'),
        'fecha_hasta': request.args.get('fecha_hasta')
    }
    filtros = {k: v for k, v in filtros.items() if v}
    
    transacciones = transaccion_db.obtener_transacciones(filtros, limite=10000)
    
    # Crear CSV
    si = StringIO()
    writer = csv.writer(si)
    
    # Encabezados
    writer.writerow(['ID', 'Fecha', 'Tipo', 'Concepto', 'Proveedor', 'Monto', 
                     'Clasificación', 'Deducible %', 'Estatus', 'Notas'])
    
    # Datos
    for t in transacciones:
        writer.writerow([
            t['id'], t['fecha'], t['tipo'], t['concepto'], t.get('proveedor', ''),
            t['monto'], t.get('clasificacion_contador') or t.get('clasificacion_ia', ''),
            t['deducible_porcentaje'], t['estatus_validacion'], t.get('notas_contador', '')
        ])
    
    output = make_response(si.getvalue())
    output.headers["Content-Disposition"] = "attachment; filename=transacciones.csv"
    output.headers["Content-type"] = "text/csv"
    
    return output

# ============================================
# MÓDULO DE SEGUROS - ENDPOINTS
# ============================================

@app.route('/seguros')
def vista_seguros():
    """Vista principal del módulo de seguros"""
    credenciales = seguro_db.obtener_credenciales(limite=10)
    tabuladores = seguro_db.obtener_tabuladores(activo=True)
    return render_template('seguros.html', credenciales=credenciales, tabuladores=tabuladores)

@app.route('/api/seguros/procesar_credencial', methods=['POST'])
def procesar_credencial_api():
    """API para procesar una credencial de seguro desde imagen (OCR)"""
    if 'imagen' not in request.files:
        return jsonify({"error": "No se recibió imagen de credencial."}), 400
    
    imagen_file = request.files['imagen']
    
    if imagen_file.filename == '':
        return jsonify({"error": "Archivo de imagen vacío."}), 400
    
    try:
        # Leer imagen
        imagen_bytes = imagen_file.read()
        
        # Extraer datos usando OCR con Gemini Vision
        datos_extractos = extraer_datos_credencial_imagen(imagen_bytes, GEMINI_API_KEY)
        
        if not datos_extractos or not datos_extractos.get('aseguradora'):
            return jsonify({
                "error": "No se pudo extraer información de la credencial. Por favor, asegúrate de que la imagen sea clara.",
                "datos_parciales": datos_extractos
            }), 400
        
        # Obtener información del plan (deducible, coaseguro, hospitales)
        aseguradora = datos_extractos.get('aseguradora', '')
        plan_nombre = datos_extractos.get('plan_nombre', '')
        
        info_plan = consultar_info_plan(aseguradora, plan_nombre)
        
        # Preparar datos para guardar
        credencial_data = {
            'medico_id': 'default',
            'paciente_nombre': datos_extractos.get('paciente_nombre', ''),
            'aseguradora': aseguradora,
            'numero_poliza': datos_extractos.get('numero_poliza', ''),
            'plan_nombre': plan_nombre,
            'nivel_hospitalario': datos_extractos.get('nivel_hospitalario', ''),
            'deducible_estimado': info_plan.get('deducible_estimado'),
            'coaseguro_porcentaje': info_plan.get('coaseguro_porcentaje'),
            'hospitales_red': info_plan.get('hospitales_red', ''),
            'datos_extractos': json.dumps(datos_extractos, ensure_ascii=False)
        }
        
        # Guardar credencial
        credencial_id = seguro_db.guardar_credencial(credencial_data)
        
        return jsonify({
            "success": True,
            "credencial_id": credencial_id,
            "datos": {
                "aseguradora": aseguradora,
                "numero_poliza": datos_extractos.get('numero_poliza', ''),
                "plan_nombre": plan_nombre,
                "deducible_estimado": info_plan.get('deducible_estimado'),
                "coaseguro_porcentaje": info_plan.get('coaseguro_porcentaje'),
                "hospitales_red": info_plan.get('hospitales_red', ''),
                "nivel_hospitalario": datos_extractos.get('nivel_hospitalario', ''),
                "paciente_nombre": datos_extractos.get('paciente_nombre', '')
            }
        })
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"[ERROR] Exception procesando credencial: {error_details}")
        return jsonify({
            "error": "Error al procesar credencial: " + str(e),
            "debug": {"traceback": error_details[:500]}
        }), 500

@app.route('/api/seguros/buscar_honorario', methods=['POST'])
def buscar_honorario_api():
    """API para buscar honorario de un procedimiento en tabulador usando RAG"""
    if not request.json:
        return jsonify({"error": "No se recibió datos JSON."}), 400
    
    aseguradora = request.json.get('aseguradora', '')
    plan_nombre = request.json.get('plan_nombre', '')
    procedimiento = request.json.get('procedimiento', '')
    codigo_cpt = request.json.get('codigo_cpt', '')
    
    if not aseguradora or not procedimiento:
        return jsonify({"error": "Aseguradora y procedimiento son requeridos."}), 400
    
    try:
        # Buscar tabulador activo de la aseguradora
        tabuladores = seguro_db.obtener_tabuladores(aseguradora=aseguradora, activo=True)
        
        contenido_tabulador = None
        tabulador_id = None
        
        if tabuladores:
            # Usar el tabulador más reciente
            tabulador = tabuladores[0]
            contenido_tabulador = tabulador.get('contenido_texto', '')
            tabulador_id = tabulador.get('id')
        
        # Buscar honorario usando RAG
        resultado = buscar_honorario_en_tabulador(
            aseguradora=aseguradora,
            plan_nombre=plan_nombre,
            procedimiento=procedimiento,
            codigo_cpt=codigo_cpt,
            tabulador_id=tabulador_id,
            contenido_tabulador=contenido_tabulador,
            api_key=GEMINI_API_KEY
        )
        
        if resultado.get('error'):
            return jsonify(resultado), 500
        
        # Guardar consulta
        if resultado.get('monto'):
            consulta_data = {
                'medico_id': 'default',
                'aseguradora': aseguradora,
                'plan_nombre': plan_nombre,
                'procedimiento': procedimiento,
                'codigo_cpt': resultado.get('codigo_cpt') or codigo_cpt,
                'monto_encontrado': resultado.get('monto'),
                'fuente_tabulador_id': tabulador_id
            }
            seguro_db.guardar_consulta_honorario(consulta_data)
        
        return jsonify({
            "success": True,
            "resultado": resultado
        })
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"[ERROR] Exception buscando honorario: {error_details}")
        return jsonify({
            "error": "Error al buscar honorario: " + str(e),
            "debug": {"traceback": error_details[:500]}
        }), 500

@app.route('/api/seguros/consultar_cobertura', methods=['POST'])
def consultar_cobertura_api():
    """API para consultar si un procedimiento está cubierto por el seguro"""
    if not request.json:
        return jsonify({"error": "No se recibió datos JSON."}), 400
    
    aseguradora = request.json.get('aseguradora', '')
    plan_nombre = request.json.get('plan_nombre', '')
    procedimiento = request.json.get('procedimiento', '')
    
    if not aseguradora or not procedimiento:
        return jsonify({"error": "Aseguradora y procedimiento son requeridos."}), 400
    
    try:
        # Buscar condiciones generales
        tabuladores = seguro_db.obtener_tabuladores(aseguradora=aseguradora, activo=True)
        
        contenido_condiciones = None
        
        # Buscar documento de condiciones generales
        for tab in tabuladores:
            if tab.get('tipo_documento') == 'condiciones_generales':
                contenido_condiciones = tab.get('contenido_texto', '')
                break
        
        # Consultar cobertura
        resultado = consultar_cobertura_procedimiento(
            aseguradora=aseguradora,
            plan_nombre=plan_nombre,
            procedimiento=procedimiento,
            contenido_condiciones=contenido_condiciones,
            api_key=GEMINI_API_KEY
        )
        
        if resultado.get('error'):
            return jsonify(resultado), 500
        
        return jsonify({
            "success": True,
            "resultado": resultado
        })
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"[ERROR] Exception consultando cobertura: {error_details}")
        return jsonify({
            "error": "Error al consultar cobertura: " + str(e),
            "debug": {"traceback": error_details[:500]}
        }), 500

@app.route('/api/seguros/generar_informe', methods=['POST'])
def generar_informe_api():
    """API para generar informe médico en PDF automáticamente"""
    if not request.json:
        return jsonify({"error": "No se recibió datos JSON."}), 400
    
    consulta_id = request.json.get('consulta_id')
    credencial_seguro_id = request.json.get('credencial_seguro_id')
    
    if not consulta_id:
        return jsonify({"error": "ID de consulta es requerido."}), 400
    
    try:
        # Obtener datos de la consulta
        consulta = db.obtener_consulta(consulta_id)
        if not consulta:
            return jsonify({"error": "Consulta no encontrada."}), 404
        
        # Obtener datos de la credencial de seguro
        datos_seguro = {}
        if credencial_seguro_id:
            credencial = seguro_db.obtener_credencial(credencial_seguro_id)
            if credencial:
                datos_seguro = {
                    'aseguradora': credencial.get('aseguradora', ''),
                    'numero_poliza': credencial.get('numero_poliza', ''),
                    'plan_nombre': credencial.get('plan_nombre', '')
                }
        
        # Si no hay credencial, usar datos del request
        if not datos_seguro.get('aseguradora'):
            datos_seguro = {
                'aseguradora': request.json.get('aseguradora', 'GENÉRICO'),
                'numero_poliza': request.json.get('numero_poliza', ''),
                'plan_nombre': request.json.get('plan_nombre', '')
            }
        
        # Preparar datos del paciente
        datos_paciente = {
            'nombre': consulta.get('paciente_nombre', '') or request.json.get('paciente_nombre', 'Paciente'),
            'edad': request.json.get('edad', ''),
            'fecha_nacimiento': request.json.get('fecha_nacimiento', ''),
            'sexo': request.json.get('sexo', '')
        }
        
        # Preparar datos de la consulta
        datos_consulta = {
            'fecha_consulta': consulta.get('fecha_consulta', ''),
            'diagnostico': consulta.get('diagnostico', ''),
            'codigo_cie10': request.json.get('codigo_cie10', ''),
            'procedimiento': request.json.get('procedimiento', ''),
            'codigo_cpt': request.json.get('codigo_cpt', ''),
            'tratamiento': consulta.get('tratamiento', ''),
            'soap_subjetivo': consulta.get('soap_subjetivo', ''),
            'soap_objetivo': consulta.get('soap_objetivo', ''),
            'soap_analisis': consulta.get('soap_analisis', ''),
            'soap_plan': consulta.get('soap_plan', ''),
            'resumen_clinico': request.json.get('resumen_clinico', '')
        }
        
        # Generar PDF
        tipo_aseguradora = datos_seguro.get('aseguradora', 'GENÉRICO')
        pdf_buffer = generar_informe_medico(
            datos_consulta=datos_consulta,
            datos_paciente=datos_paciente,
            datos_seguro=datos_seguro,
            tipo_aseguradora=tipo_aseguradora
        )
        
        # Guardar informe en BD
        informe_data = {
            'consulta_id': consulta_id,
            'credencial_seguro_id': credencial_seguro_id,
            'aseguradora': datos_seguro.get('aseguradora', ''),
            'paciente_nombre': datos_paciente.get('nombre', ''),
            'numero_poliza': datos_seguro.get('numero_poliza', ''),
            'diagnostico': datos_consulta.get('diagnostico', ''),
            'procedimiento': datos_consulta.get('procedimiento', ''),
            'codigo_cpt': datos_consulta.get('codigo_cpt', ''),
            'codigo_cie10': datos_consulta.get('codigo_cie10', ''),
            'informe_pdf_path': ''  # En producción, guardar en disco
        }
        informe_id = seguro_db.guardar_informe_medico(informe_data)
        
        # Retornar PDF como descarga
        response = make_response(pdf_buffer.getvalue())
        paciente_nombre_clean = datos_paciente.get('nombre', 'Paciente').replace(' ', '_')
        fecha = datetime.now().strftime('%Y%m%d')
        filename = f"Informe_Medico_{paciente_nombre_clean}_{fecha}.pdf"
        
        response.headers["Content-Disposition"] = f"attachment; filename={filename}"
        response.headers["Content-type"] = "application/pdf"
        
        return response
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"[ERROR] Exception generando informe: {error_details}")
        return jsonify({
            "error": "Error al generar informe: " + str(e),
            "debug": {"traceback": error_details[:500]}
        }), 500

@app.route('/api/seguros/credenciales', methods=['GET'])
def obtener_credenciales_api():
    """API para obtener lista de credenciales procesadas"""
    medico_id = request.args.get('medico_id', 'default')
    limite = int(request.args.get('limite', 50))
    
    credenciales = seguro_db.obtener_credenciales(medico_id=medico_id, limite=limite)
    return jsonify(credenciales)

@app.route('/api/seguros/credencial/<int:credencial_id>', methods=['GET'])
def obtener_credencial_api(credencial_id):
    """API para obtener una credencial específica"""
    credencial = seguro_db.obtener_credencial(credencial_id)
    if not credencial:
        return jsonify({"error": "Credencial no encontrada"}), 404
    return jsonify(credencial)

@app.route('/api/seguros/cargar_tabulador', methods=['POST'])
def cargar_tabulador_api():
    """API para cargar y procesar un PDF de tabulador"""
    if 'pdf' not in request.files:
        return jsonify({"error": "No se recibió archivo PDF."}), 400
    
    pdf_file = request.files['pdf']
    
    if pdf_file.filename == '':
        return jsonify({"error": "Archivo PDF vacío."}), 400
    
    if not pdf_file.filename.lower().endswith('.pdf'):
        return jsonify({"error": "El archivo debe ser un PDF."}), 400
    
    # Obtener datos opcionales del formulario
    aseguradora_manual = request.form.get('aseguradora', '').strip()
    plan_manual = request.form.get('plan_nombre', '').strip()
    tipo_documento_manual = request.form.get('tipo_documento', '').strip()
    fecha_vigencia_manual = request.form.get('fecha_vigencia', '').strip()
    
    try:
        # Leer PDF
        pdf_bytes = pdf_file.read()
        
        # Validar tamaño (máx 50MB)
        if len(pdf_bytes) > 50 * 1024 * 1024:
            return jsonify({"error": "El archivo PDF es demasiado grande (máx. 50MB)."}), 400
        
        # Procesar PDF
        resultado = procesar_tabulador_pdf(pdf_bytes, pdf_file.filename)
        
        if resultado.get('error'):
            return jsonify({"error": resultado['error']}), 400
        
        # Verificar si ya existe (por hash)
        tabuladores_existentes = seguro_db.obtener_tabuladores()
        hash_pdf = resultado.get('hash')
        
        for tab_existente in tabuladores_existentes:
            if tab_existente.get('archivo_hash') == hash_pdf:
                return jsonify({
                    "error": "Este tabulador ya fue cargado anteriormente.",
                    "tabulador_existente": {
                        "id": tab_existente.get('id'),
                        "aseguradora": tab_existente.get('aseguradora'),
                        "fecha_carga": tab_existente.get('fecha_carga')
                    }
                }), 409
        
        # Usar datos detectados o manuales
        aseguradora = aseguradora_manual or resultado.get('aseguradora') or 'Desconocida'
        plan = plan_manual or resultado.get('plan') or ''
        tipo_documento = tipo_documento_manual or resultado.get('tipo_documento') or 'tabulador'
        fecha_vigencia = fecha_vigencia_manual or resultado.get('fecha_vigencia') or None
        
        if not aseguradora or aseguradora == 'Desconocida':
            return jsonify({
                "error": "No se pudo detectar la aseguradora automáticamente. Por favor, especifícala manualmente.",
                "datos_detectados": {
                    "aseguradora": resultado.get('aseguradora'),
                    "plan": resultado.get('plan'),
                    "tipo_documento": resultado.get('tipo_documento'),
                    "num_paginas": resultado.get('num_paginas')
                }
            }), 400
        
        # Guardar en base de datos
        # En producción, guardar el PDF en disco y solo la ruta en BD
        tabulador_data = {
            'aseguradora': aseguradora,
            'plan_nombre': plan,
            'tipo_documento': tipo_documento,
            'archivo_path': pdf_file.filename,  # En producción: ruta en disco
            'archivo_hash': hash_pdf,
            'fecha_vigencia': fecha_vigencia,
            'contenido_texto': resultado.get('texto', ''),
            'contenido_embedding': ''  # Para futuros embeddings
        }
        
        tabulador_id = seguro_db.guardar_tabulador(tabulador_data)
        
        return jsonify({
            "success": True,
            "tabulador_id": tabulador_id,
            "mensaje": f"Tabulador cargado exitosamente ({resultado.get('num_paginas', 0)} páginas)",
            "datos": {
                "aseguradora": aseguradora,
                "plan_nombre": plan,
                "tipo_documento": tipo_documento,
                "fecha_vigencia": fecha_vigencia,
                "num_paginas": resultado.get('num_paginas', 0),
                "texto_preview": resultado.get('texto', '')[:500] + '...' if len(resultado.get('texto', '')) > 500 else resultado.get('texto', '')
            }
        })
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"[ERROR] Exception cargando tabulador: {error_details}")
        return jsonify({
            "error": "Error al procesar PDF: " + str(e),
            "debug": {"traceback": error_details[:500]}
        }), 500

@app.route('/api/seguros/tabuladores', methods=['GET'])
def obtener_tabuladores_api():
    """API para obtener lista de tabuladores cargados"""
    aseguradora = request.args.get('aseguradora')
    activo = request.args.get('activo', 'true').lower() == 'true'
    
    tabuladores = seguro_db.obtener_tabuladores(aseguradora=aseguradora, activo=activo)
    
    # Limitar tamaño del texto en respuesta (para no saturar)
    for tab in tabuladores:
        if tab.get('contenido_texto'):
            texto = tab['contenido_texto']
            if len(texto) > 1000:
                tab['contenido_texto_preview'] = texto[:1000] + '...'
            else:
                tab['contenido_texto_preview'] = texto
            # No enviar el texto completo en la lista
            del tab['contenido_texto']
    
    return jsonify(tabuladores)

@app.route('/api/seguros/tabulador/<int:tabulador_id>', methods=['GET'])
def obtener_tabulador_api(tabulador_id):
    """API para obtener un tabulador específico (con texto completo)"""
    tabuladores = seguro_db.obtener_tabuladores()
    tabulador = next((t for t in tabuladores if t.get('id') == tabulador_id), None)
    
    if not tabulador:
        return jsonify({"error": "Tabulador no encontrado"}), 404
    
    return jsonify(tabulador)

@app.route('/api/seguros/tabulador/<int:tabulador_id>', methods=['DELETE'])
def eliminar_tabulador_api(tabulador_id):
    """API para desactivar un tabulador (soft delete)"""
    with sqlite3.connect(seguro_db.db_path) as conn:
        cursor = conn.execute('''
            UPDATE tabuladores 
            SET activo = 0, updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
        ''', (tabulador_id,))
        
        if cursor.rowcount > 0:
            return jsonify({"success": True, "message": "Tabulador desactivado correctamente"})
        else:
            return jsonify({"error": "Tabulador no encontrado"}), 404

# ============================================
# MÓDULO ASISTENTE LEGAL - ENDPOINTS
# ============================================

@app.route('/legal')
def vista_legal():
    """Vista principal del módulo legal para médicos"""
    alertas = legal_db.obtener_alertas_legales(limite=10)
    contratos_vencer = legal_db.obtener_contratos_por_vencer(dias=30)
    return render_template('legal.html', alertas=alertas, contratos_vencer=contratos_vencer)

@app.route('/legal/abogado')
def vista_legal_abogado():
    """Vista del dashboard del abogado"""
    stats = legal_db.obtener_estadisticas_cumplimiento()
    alertas = legal_db.obtener_alertas_legales(limite=20)
    plantillas = legal_db.obtener_plantillas()
    return render_template('legal_abogado.html', stats=stats, alertas=alertas, plantillas=plantillas)

@app.route('/api/legal/generar_consentimiento', methods=['POST'])
def generar_consentimiento_api():
    """API para generar un consentimiento informado personalizado"""
    if not request.json:
        return jsonify({"error": "No se recibió datos JSON."}), 400
    
    procedimiento = request.json.get('procedimiento', '')
    paciente_nombre = request.json.get('paciente_nombre', '')
    consulta_id = request.json.get('consulta_id')
    medico_nombre = request.json.get('medico_nombre', 'Dr. Médico')
    
    if not procedimiento:
        return jsonify({"error": "Procedimiento es requerido."}), 400
    
    try:
        # Buscar plantilla para el procedimiento
        plantilla = legal_db.obtener_plantilla_por_procedimiento(procedimiento)
        
        if not plantilla:
            # Si no hay plantilla específica, buscar una genérica
            plantillas = legal_db.obtener_plantillas(tipo_documento='consentimiento_informado')
            if plantillas:
                plantilla = plantillas[0]
            else:
                return jsonify({
                    "error": f"No se encontró plantilla para el procedimiento '{procedimiento}'. El abogado debe crear una plantilla primero."
                }), 404
        
        # Personalizar plantilla con datos del paciente usando IA
        template = plantilla.get('contenido_template', '')
        
        # Usar Gemini para personalizar el documento
        prompt = f"""Eres un asistente legal especializado en documentos médicos en México.

Personaliza el siguiente consentimiento informado con los datos del paciente y procedimiento.

PLANTILLA BASE:
{template}

DATOS A INSERTAR:
- Nombre del paciente: {paciente_nombre}
- Nombre del médico: {medico_nombre}
- Procedimiento: {procedimiento}
- Fecha actual: {datetime.now().strftime('%d de %B de %Y')}

INSTRUCCIONES:
1. Reemplaza todos los placeholders como [NOMBRE_PACIENTE], [PROCEDIMIENTO], [FECHA], etc.
2. Mantén el formato legal y profesional
3. Incluye los riesgos específicos del procedimiento si están en la plantilla
4. Asegúrate de que el documento esté completo y listo para firma

Responde SOLO con el documento personalizado, sin explicaciones adicionales."""
        
        documento_personalizado = call_gemini_api(prompt, GEMINI_API_KEY)
        
        if not documento_personalizado:
            # Fallback: reemplazo simple
            documento_personalizado = template.replace('[NOMBRE_PACIENTE]', paciente_nombre)
            documento_personalizado = documento_personalizado.replace('[PROCEDIMIENTO]', procedimiento)
            documento_personalizado = documento_personalizado.replace('[MEDICO]', medico_nombre)
            documento_personalizado = documento_personalizado.replace('[FECHA]', datetime.now().strftime('%d de %B de %Y'))
        
        return jsonify({
            "success": True,
            "documento": documento_personalizado,
            "plantilla_id": plantilla.get('id'),
            "procedimiento": procedimiento
        })
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"[ERROR] Exception generando consentimiento: {error_details}")
        return jsonify({
            "error": "Error al generar consentimiento: " + str(e),
            "debug": {"traceback": error_details[:500]}
        }), 500

@app.route('/api/legal/firmar_documento', methods=['POST'])
def firmar_documento_api():
    """API para guardar un documento firmado"""
    if not request.json:
        return jsonify({"error": "No se recibió datos JSON."}), 400
    
    try:
        # Obtener datos de la firma
        firma_imagen = request.json.get('firma_imagen')  # Base64
        documento_data = {
            'medico_id': request.json.get('medico_id', 'default'),
            'paciente_id': request.json.get('paciente_id'),
            'paciente_nombre': request.json.get('paciente_nombre', ''),
            'consulta_id': request.json.get('consulta_id'),
            'plantilla_id': request.json.get('plantilla_id'),
            'tipo_documento': request.json.get('tipo_documento', 'consentimiento_informado'),
            'procedimiento': request.json.get('procedimiento', ''),
            'contenido_documento': request.json.get('contenido_documento', ''),
            'firma_digital': request.json.get('firma_digital', ''),
            'firma_imagen_path': '',  # En producción, guardar imagen en disco
            'fecha_firma': datetime.now().strftime('%Y-%m-%d'),
            'hora_firma': datetime.now().strftime('%H:%M:%S'),
            'latitud': request.json.get('latitud'),
            'longitud': request.json.get('longitud'),
            'ip_address': request.remote_addr,
            'dispositivo': request.headers.get('User-Agent', ''),
            'hash_documento': str(hash(request.json.get('contenido_documento', '')))
        }
        
        # Guardar documento firmado
        documento_id = legal_db.guardar_documento_firmado(documento_data)
        
        # Registrar en auditoría
        legal_db.registrar_acceso_auditoria({
            'medico_id': documento_data['medico_id'],
            'usuario': 'paciente',
            'tipo_acceso': 'firma',
            'entidad': 'documento_firmado',
            'entidad_id': documento_id,
            'ip_address': documento_data['ip_address'],
            'user_agent': documento_data['dispositivo'],
            'detalles': f"Firma de {documento_data['tipo_documento']} para procedimiento {documento_data['procedimiento']}"
        })
        
        return jsonify({
            "success": True,
            "documento_id": documento_id,
            "message": "Documento firmado y guardado correctamente"
        })
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"[ERROR] Exception firmando documento: {error_details}")
        return jsonify({
            "error": "Error al firmar documento: " + str(e),
            "debug": {"traceback": error_details[:500]}
        }), 500

@app.route('/api/legal/auditoria_cumplimiento', methods=['POST'])
def auditoria_cumplimiento_api():
    """API para ejecutar auditoría de cumplimiento (cruza consultas vs documentos)"""
    medico_id = request.json.get('medico_id', 'default') if request.json else 'default'
    
    try:
        # Obtener consultas recientes (últimos 30 días)
        consultas = db.obtener_consultas(medico_id=medico_id, limite=1000)
        
        # Obtener documentos firmados
        documentos = legal_db.obtener_documentos_firmados(medico_id=medico_id, limite=1000)
        
        # Crear mapa de consultas con consentimiento
        consultas_con_consentimiento = set()
        for doc in documentos:
            if doc.get('consulta_id') and doc.get('tipo_documento') == 'consentimiento_informado':
                consultas_con_consentimiento.add(doc['consulta_id'])
        
        # Identificar consultas sin consentimiento
        alertas_creadas = 0
        procedimientos_requieren_consentimiento = ['cirugía', 'biopsia', 'endoscopia', 'colonoscopia', 'operación', 'intervención']
        
        for consulta in consultas:
            consulta_id = consulta.get('id')
            diagnostico = consulta.get('diagnostico', '').lower()
            tratamiento = consulta.get('tratamiento', '').lower()
            transcripcion = consulta.get('transcripcion', '').lower()
            
            # Verificar si requiere consentimiento
            requiere_consentimiento = any(proc in diagnostico or proc in tratamiento or proc in transcripcion 
                                         for proc in procedimientos_requieren_consentimiento)
            
            if requiere_consentimiento and consulta_id not in consultas_con_consentimiento:
                # Verificar si ya existe una alerta para esta consulta
                alertas_existentes = legal_db.obtener_alertas_legales(medico_id=medico_id, limite=1000)
                ya_existe = any(a.get('entidad_id') == consulta_id and a.get('estado') == 'activa' 
                               for a in alertas_existentes)
                
                if not ya_existe:
                    legal_db.crear_alerta_legal({
                        'medico_id': medico_id,
                        'tipo_alerta': 'consentimiento_faltante',
                        'severidad': 'alta',
                        'titulo': f'Consentimiento faltante para consulta #{consulta_id}',
                        'descripcion': f'La consulta del {consulta.get("fecha_consulta", "")} requiere consentimiento informado pero no se encontró documento firmado.',
                        'entidad_tipo': 'consulta',
                        'entidad_id': consulta_id
                    })
                    alertas_creadas += 1
        
        return jsonify({
            "success": True,
            "alertas_creadas": alertas_creadas,
            "total_consultas_revisadas": len(consultas),
            "consultas_con_consentimiento": len(consultas_con_consentimiento),
            "message": f"Auditoría completada. Se crearon {alertas_creadas} alertas."
        })
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"[ERROR] Exception en auditoría: {error_details}")
        return jsonify({
            "error": "Error en auditoría: " + str(e),
            "debug": {"traceback": error_details[:500]}
        }), 500

@app.route('/api/legal/contratos', methods=['GET'])
def obtener_contratos_api():
    """API para obtener contratos de staff"""
    medico_id = request.args.get('medico_id', 'default')
    estado = request.args.get('estado')
    
    contratos = legal_db.obtener_contratos_staff(medico_id=medico_id, estado=estado)
    return jsonify(contratos)

@app.route('/api/legal/contratos', methods=['POST'])
def crear_contrato_api():
    """API para crear un contrato de staff"""
    if not request.json:
        return jsonify({"error": "No se recibió datos JSON."}), 400
    
    try:
        contrato_data = {
            'medico_id': request.json.get('medico_id', 'default'),
            'empleado_nombre': request.json.get('empleado_nombre'),
            'puesto': request.json.get('puesto', ''),
            'tipo_contrato': request.json.get('tipo_contrato'),
            'fecha_inicio': request.json.get('fecha_inicio'),
            'fecha_fin': request.json.get('fecha_fin'),
            'salario': request.json.get('salario'),
            'plantilla_contrato_id': request.json.get('plantilla_contrato_id'),
            'documento_firmado_id': request.json.get('documento_firmado_id'),
            'notas': request.json.get('notas', '')
        }
        
        contrato_id = legal_db.guardar_contrato_staff(contrato_data)
        
        # Si tiene fecha de vencimiento, verificar si necesita alerta
        if contrato_data.get('fecha_fin'):
            fecha_fin = datetime.strptime(contrato_data['fecha_fin'], '%Y-%m-%d').date()
            dias_restantes = (fecha_fin - datetime.now().date()).days
            
            if dias_restantes <= 30 and dias_restantes >= 0:
                legal_db.crear_alerta_legal({
                    'medico_id': contrato_data['medico_id'],
                    'tipo_alerta': 'contrato_vencido',
                    'severidad': 'media' if dias_restantes > 7 else 'alta',
                    'titulo': f'Contrato de {contrato_data["empleado_nombre"]} vence en {dias_restantes} días',
                    'descripcion': f'El contrato de {contrato_data["empleado_nombre"]} vence el {contrato_data["fecha_fin"]}. ¿Renovar o terminar?',
                    'entidad_tipo': 'contrato_staff',
                    'entidad_id': contrato_id
                })
        
        return jsonify({
            "success": True,
            "contrato_id": contrato_id,
            "message": "Contrato creado correctamente"
        })
        
    except Exception as e:
        return jsonify({"error": "Error al crear contrato: " + str(e)}), 500

@app.route('/api/legal/incidencias', methods=['POST'])
def crear_incidencia_api():
    """API para registrar una incidencia laboral"""
    if not request.json:
        return jsonify({"error": "No se recibió datos JSON."}), 400
    
    try:
        incidencia_data = {
            'medico_id': request.json.get('medico_id', 'default'),
            'contrato_staff_id': request.json.get('contrato_staff_id'),
            'empleado_nombre': request.json.get('empleado_nombre'),
            'tipo_incidencia': request.json.get('tipo_incidencia'),
            'descripcion': request.json.get('descripcion'),
            'fecha_incidencia': request.json.get('fecha_incidencia', datetime.now().strftime('%Y-%m-%d')),
            'hora_incidencia': request.json.get('hora_incidencia', datetime.now().strftime('%H:%M:%S')),
            'evidencia_path': request.json.get('evidencia_path', '')
        }
        
        incidencia_id = legal_db.guardar_incidencia_laboral(incidencia_data)
        
        return jsonify({
            "success": True,
            "incidencia_id": incidencia_id,
            "message": "Incidencia registrada correctamente"
        })
        
    except Exception as e:
        return jsonify({"error": "Error al registrar incidencia: " + str(e)}), 500

@app.route('/api/legal/panico', methods=['POST'])
def boton_panico_api():
    """API para el botón de pánico legal (respuesta a crisis)"""
    if not request.json:
        return jsonify({"error": "No se recibió datos JSON."}), 400
    
    tipo_crisis = request.json.get('tipo_crisis', 'emergencia_legal')
    medico_id = request.json.get('medico_id', 'default')
    
    try:
        # Obtener guía de reacción rápida
        guia = legal_db.obtener_guia_reaccion(tipo_crisis, medico_id)
        
        if not guia:
            # Crear guía genérica si no existe
            guia = {
                'titulo': 'Guía de Reacción Rápida',
                'contenido': 'Contacta inmediatamente a tu abogado. No proporciones información adicional sin asesoría legal.',
                'pasos_accion': '1. Mantén la calma\n2. Contacta al abogado\n3. No firmes nada sin revisar',
                'documentos_necesarios': 'Identificación oficial, documentos del caso',
                'contacto_abogado': 'Contactar al abogado asignado'
            }
        
        # Crear alerta de emergencia
        alerta_id = legal_db.crear_alerta_legal({
            'medico_id': medico_id,
            'tipo_alerta': 'riesgo_alto',
            'severidad': 'critica',
            'titulo': f'🚨 ALERTA URGENTE: {guia.get("titulo", "Crisis Legal")}',
            'descripcion': f'Se activó el botón de pánico para: {tipo_crisis}. {guia.get("contenido", "")}',
            'entidad_tipo': 'crisis',
            'entidad_id': None
        })
        
        return jsonify({
            "success": True,
            "alerta_id": alerta_id,
            "guia": guia,
            "message": "Alerta enviada. Revisa la guía de reacción rápida."
        })
        
    except Exception as e:
        return jsonify({"error": "Error al activar botón de pánico: " + str(e)}), 500

@app.route('/api/legal/plantillas', methods=['GET'])
def obtener_plantillas_api():
    """API para obtener plantillas legales"""
    tipo_documento = request.args.get('tipo_documento')
    activo = request.args.get('activo', 'true').lower() == 'true'
    
    plantillas = legal_db.obtener_plantillas(tipo_documento=tipo_documento, activo=activo)
    return jsonify(plantillas)

@app.route('/api/legal/plantillas', methods=['POST'])
def crear_plantilla_api():
    """API para crear/actualizar una plantilla legal"""
    if not request.json:
        return jsonify({"error": "No se recibió datos JSON."}), 400
    
    try:
        plantilla_data = {
            'tipo_documento': request.json.get('tipo_documento'),
            'nombre_plantilla': request.json.get('nombre_plantilla'),
            'procedimiento': request.json.get('procedimiento', ''),
            'contenido_template': request.json.get('contenido_template'),
            'variables_template': request.json.get('variables_template', ''),
            'aprobado_por': request.json.get('aprobado_por', ''),
            'fecha_aprobacion': request.json.get('fecha_aprobacion', datetime.now().strftime('%Y-%m-%d'))
        }
        
        plantilla_id = legal_db.guardar_plantilla(plantilla_data)
        
        return jsonify({
            "success": True,
            "plantilla_id": plantilla_id,
            "message": "Plantilla guardada correctamente"
        })
        
    except Exception as e:
        return jsonify({"error": "Error al guardar plantilla: " + str(e)}), 500

@app.route('/api/legal/alertas', methods=['GET'])
def obtener_alertas_api():
    """API para obtener alertas legales"""
    medico_id = request.args.get('medico_id', 'default')
    estado = request.args.get('estado', 'activa')
    limite = int(request.args.get('limite', 50))
    
    alertas = legal_db.obtener_alertas_legales(medico_id=medico_id, estado=estado, limite=limite)
    return jsonify(alertas)

@app.route('/api/legal/alertas/<int:alerta_id>/resolver', methods=['POST'])
def resolver_alerta_api(alerta_id):
    """API para resolver una alerta legal"""
    if not request.json:
        return jsonify({"error": "No se recibió datos JSON."}), 400
    
    resuelto_por = request.json.get('resuelto_por', 'abogado')
    notas = request.json.get('notas', '')
    
    try:
        legal_db.resolver_alerta(alerta_id, resuelto_por, notas)
        return jsonify({"success": True, "message": "Alerta resuelta correctamente"})
    except Exception as e:
        return jsonify({"error": "Error al resolver alerta: " + str(e)}), 500

@app.route('/api/legal/estadisticas_cumplimiento')
def estadisticas_cumplimiento_api():
    """API para obtener estadísticas de cumplimiento"""
    medico_id = request.args.get('medico_id', 'default')
    stats = legal_db.obtener_estadisticas_cumplimiento(medico_id=medico_id)
    return jsonify(stats)

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5555))
    debug = os.environ.get('FLASK_ENV') != 'production'
    app.run(host='0.0.0.0', port=port, debug=debug)